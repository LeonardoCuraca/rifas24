import openpyxl
from django import template
from django.http import HttpResponse

register = template.Library()

@register.simple_tag
def export_to_excel(response, model):
    headers = []
    for field in model._meta.fields:
        headers.append(field.verbose_name.title())

    data = model.objects.all().values_list(*headers)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model._meta.verbose_name_plural.title()

    # Write header row
    for col_num, header_title in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws[f'{col_letter}1'] = header_title

    # Write data rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            ws[f'{col_letter}{row_num}'] = cell_value

    wb.save(response)
    return response